import pandas as pd
import os
from openai import OpenAI
from dotenv import load_dotenv
import json
import io
import openpyxl
import csv
from datetime import datetime

# Load environment variables
load_dotenv()

class GPTFinancialParser:
    """
    A flexible parser that uses OpenAI's GPT to transform unformatted financial documents
    into structured tables with New Zealand-specific financial categorization.
    """
    
    def __init__(self):
        """Initialize the parser with OpenAI API key from environment or secrets."""
        # Get API key from environment variable or Streamlit secrets
        api_key = os.getenv("OPENAI_API_KEY")
        
        # For Streamlit Cloud, try to get from st.secrets
        if not api_key:
            try:
                import streamlit as st
                if hasattr(st, 'secrets') and "OPENAI_API_KEY" in st.secrets:
                    api_key = st.secrets["OPENAI_API_KEY"]
            except:
                pass
        
        if not api_key:
            raise ValueError("OpenAI API key not found. Please set the OPENAI_API_KEY environment variable.")
        
        # Initialize OpenAI client
        self.client = OpenAI(api_key=api_key)
    
    def extract_file_content(self, file_path):
        """Extract content from file based on its extension."""
        _, ext = os.path.splitext(file_path)
        ext = ext.lower()
        
        if ext in ['.xlsx', '.xls']:
            return self._extract_excel_content(file_path)
        elif ext == '.csv':
            return self._extract_csv_content(file_path)
        elif ext == '.pdf':
            return self._extract_pdf_content(file_path)
        elif ext in ['.jpg', '.jpeg', '.png']:
            return self._extract_image_content(file_path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")
    
    def _extract_excel_content(self, file_path):
        """Extract content from Excel file."""
        try:
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Convert to string representation
            buffer = io.StringIO()
            df.to_csv(buffer, index=False)
            content = buffer.getvalue()
            
            # Also get sheet names and first few rows of each sheet for context
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_info = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = []
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i >= 10:  # Get first 10 rows
                        break
                    rows.append(row)
                sheet_info.append({
                    "name": sheet_name,
                    "sample_rows": rows
                })
            
            return {
                "main_content": content,
                "sheet_info": sheet_info
            }
        except Exception as e:
            raise ValueError(f"Error extracting Excel content: {e}")
    
    def _extract_csv_content(self, file_path):
        """Extract content from CSV file."""
        try:
            with open(file_path, 'r') as f:
                content = f.read()
            return {"main_content": content}
        except Exception as e:
            raise ValueError(f"Error extracting CSV content: {e}")
    
    def _extract_pdf_content(self, file_path):
        """Extract content from PDF file."""
        try:
            import pdfplumber
            
            content = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    content += page.extract_text() + "\n\n"
                    
                    # Extract tables if any
                    tables = page.extract_tables()
                    if tables:
                        content += "TABLES:\n"
                        for table in tables:
                            for row in table:
                                content += ",".join([str(cell) if cell else "" for cell in row]) + "\n"
                            content += "\n"
            
            return {"main_content": content}
        except Exception as e:
            raise ValueError(f"Error extracting PDF content: {e}")
    
    def _extract_image_content(self, file_path):
        """Extract content from image file using OCR."""
        try:
            import pytesseract
            from PIL import Image
            
            img = Image.open(file_path)
            content = pytesseract.image_to_string(img)
            
            return {"main_content": content}
        except Exception as e:
            raise ValueError(f"Error extracting image content: {e}")
    
    def transform_to_structured_table(self, content):
        """
        Use GPT to transform unformatted financial document content into a structured table
        with New Zealand-specific financial categorization.
        
        Args:
            content (str): Raw content from the financial document
            
        Returns:
            pandas.DataFrame: Structured table with standardized columns for NZ financial data
        """
        try:
            # User-provided prompt for NZ-specific financial data extraction
            prompt = f"""
            Your task is to read the uploaded financial file — which may contain Profit and Loss Statements, Balance Sheets, Cashflow Statements, or financial transactions — and consolidate all available financial data into a single structured JSON table with these fields:
            • HighLevelCategory (Revenue, Cost of Goods Sold, Operating Expenses, Assets, Liabilities, Equity, Cashflow, GST, Other)
            • Subcategory (Specific subcategory name like Sales Revenue, Donations, Accounts Payable)
            • Amount (numeric, no dollar signs, no commas)
            • Entity (Department or Team Name if available; otherwise null)
            • Period (e.g., 'March 2025', 'FY2024 Q1') if available; otherwise null
            • Date (specific transaction date if available, otherwise null)
            • GST_Treatment (Standard, Zero-Rated, Exempt, or Unknown)
            • Currency (NZD)

            Instructions:
            • Consolidate multiple departments, time periods, and sheets if available.
            • Categorise all line items according to New Zealand accounting practices.
            • Always use New Zealand English spelling (e.g., categorise, organisation).
            • If values are missing, set fields to null.
            • Assume all amounts are in NZD unless stated otherwise.
            • Only return strict JSON. No explanations, no commentary.

            Return ONLY a JSON object with the format:
            {{
                "table_data": [
                    {{
                        "HighLevelCategory": "Category name",
                        "Subcategory": "Subcategory name",
                        "Amount": 1234.56,
                        "Entity": "Department name or null",
                        "Period": "Period description or null",
                        "Date": "YYYY-MM-DD or null",
                        "GST_Treatment": "Standard/Zero-Rated/Exempt/Unknown",
                        "Currency": "NZD"
                    }},
                    ...more rows...
                ]
            }}

            Document content:
            {content[:8000]}  # Limit content to avoid token limits
            """
            
            print("Sending content to GPT for transformation with NZ-specific financial categorization...")
            
            response = self.client.chat.completions.create(
                model="gpt-4-turbo",
                messages=[
                    {"role": "system", "content": "You are a financial data extraction expert specializing in New Zealand accounting practices."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            
            result = json.loads(response.choices[0].message.content)
            print(f"GPT transformation complete. Extracted {len(result.get('table_data', []))} rows of data with NZ-specific categorization.")
            
            # Convert to DataFrame
            df = pd.DataFrame(result.get("table_data", []))
            
            # Ensure correct data types
            if 'Amount' in df.columns:
                df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
            
            if 'Date' in df.columns:
                df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.strftime('%Y-%m-%d')
            
            # Map the new columns to the expected columns in the application
            column_mapping = {
                'HighLevelCategory': 'main_category',
                'Subcategory': 'subcategory',
                'Amount': 'amount',
                'Date': 'date'
            }
            
            # Create a new DataFrame with the expected columns
            app_df = pd.DataFrame()
            
            # Map columns and fill with defaults if missing
            for new_col, app_col in column_mapping.items():
                if new_col in df.columns:
                    app_df[app_col] = df[new_col]
                else:
                    if app_col == 'date':
                        app_df[app_col] = pd.Timestamp.now().strftime('%Y-%m-%d')
                    elif app_col == 'amount':
                        app_df[app_col] = 0.0
                    elif app_col == 'main_category':
                        app_df[app_col] = "Uncategorized"
                    elif app_col == 'subcategory':
                        app_df[app_col] = "Unknown"
            
            # Create description field from available information
            if 'Entity' in df.columns and 'Period' in df.columns:
                # Combine Entity and Period for description if available
                app_df['description'] = df.apply(
                    lambda row: f"{row['Entity'] if pd.notna(row['Entity']) else ''} - "
                                f"{row['Subcategory']} "
                                f"({row['Period']})" if pd.notna(row['Period']) else f"{row['Subcategory']}",
                    axis=1
                )
            elif 'Subcategory' in df.columns:
                app_df['description'] = df['Subcategory']
            else:
                app_df['description'] = "Unknown"
            
            # Store the original data with all fields for potential future use
            app_df['original_data'] = df.apply(lambda x: x.to_dict(), axis=1)
            
            # Ensure we have all required columns for the application
            required_columns = ['date', 'description', 'amount', 'main_category', 'subcategory']
            for col in required_columns:
                if col not in app_df.columns:
                    if col == 'description':
                        app_df['description'] = "Unknown"
                    elif col == 'date':
                        app_df['date'] = pd.Timestamp.now().strftime('%Y-%m-%d')
                    elif col == 'amount':
                        app_df['amount'] = 0.0
                    elif col == 'main_category':
                        app_df['main_category'] = "Uncategorized"
                    elif col == 'subcategory':
                        app_df['subcategory'] = "Unknown"
            
            # Return the structured table with the required columns first, then any additional columns
            return app_df[required_columns + [col for col in app_df.columns if col not in required_columns]]
        
        except Exception as e:
            print(f"Error transforming document to structured table: {e}")
            # Return empty DataFrame with required columns
            return pd.DataFrame(columns=['date', 'description', 'amount', 'main_category', 'subcategory'])
    
    def parse_financial_document(self, file_path):
        """
        Parse a financial document using GPT to transform it into a structured table with NZ-specific categorization.
        
        Args:
            file_path (str): Path to the financial document file
            
        Returns:
            pandas.DataFrame: DataFrame with structured financial data for NZ accounting
        """
        try:
            # Extract content from file
            content_data = self.extract_file_content(file_path)
            content = content_data["main_content"]
            
            # Transform content to structured table with NZ-specific categorization
            df = self.transform_to_structured_table(content)
            
            return df
        except Exception as e:
            print(f"Error parsing financial document: {e}")
            # Return empty DataFrame with standard columns
            return pd.DataFrame(columns=['date', 'description', 'amount', 'main_category', 'subcategory'])

# Function to use in the main application
def parse_with_gpt(file_path):
    """
    Parse a financial document using GPT to transform it into a structured table with NZ-specific categorization.
    
    Args:
        file_path (str): Path to the financial document file
        
    Returns:
        pandas.DataFrame: DataFrame with structured financial data for NZ accounting
    """
    try:
        parser = GPTFinancialParser()
        return parser.parse_financial_document(file_path)
    except Exception as e:
        print(f"Error parsing with GPT: {e}")
        return pd.DataFrame(columns=['date', 'description', 'amount', 'main_category', 'subcategory'])
